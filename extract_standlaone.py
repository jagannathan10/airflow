#!/usr/bin/env python3
import json
import csv
import argparse
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


def as_list(x):
    if x is None:
        return []
    return x if isinstance(x, list) else [x]


def join_args(args) -> str:
    return " ".join(str(a) for a in as_list(args) if str(a).strip())


def build_command(job: Dict[str, Any]) -> str:
    # Control-M export format in your sample: FilePath + FileName + optional Arguments
    fp = (job.get("FilePath") or "").strip()
    fn = (job.get("FileName") or "").strip()
    args = join_args(job.get("Arguments"))

    if not fp or not fn:
        return ""

    cmd = f"{fp.rstrip('/')}/{fn}"
    if args:
        cmd += f" {args}"
    return cmd


def schedule_to_text(when: Any) -> str:
    # Convert When{} to one readable string; return "" if missing
    if not isinstance(when, dict):
        return ""

    wdays = ",".join(as_list(when.get("WeekDays")))
    from_t = (when.get("FromTime") or "").strip()
    to_t = (when.get("ToTime") or "").strip()
    cal = (when.get("MonthDaysCalendar") or "").strip()

    parts = []
    if wdays:
        parts.append(f"WeekDays={wdays}")
    if from_t:
        parts.append(f"FromTime={from_t}")
    if to_t:
        parts.append(f"ToTime={to_t}")
    if cal:
        parts.append(f"Calendar={cal}")

    return " ; ".join(parts)


def has_calendar(when: Any) -> bool:
    return isinstance(when, dict) and bool((when.get("MonthDaysCalendar") or "").strip())


def is_filewatcher(job: Dict[str, Any]) -> bool:
    # Your sample uses fileWatcher.ksh for event-driven jobs
    return (job.get("FileName") or "").strip().lower() == "filewatcher.ksh"


def is_workflow_dep(job: Dict[str, Any]) -> bool:
    # In your sample, workflow chaining is modeled as eventsToWaitFor
    return bool(job.get("eventsToWaitFor"))


def is_candidate_job(job: Dict[str, Any]) -> bool:
    """
    Candidate for "standalone/singleton list":
    - must be Job:Script
    - must have runnable script info (FilePath/FileName)
    - exclude workflow deps (eventsToWaitFor)
    - exclude event-driven filewatcher jobs
    NOTE: We allow missing schedule so we can highlight those rows in red.
    """
    if job.get("Type") != "Job:Script":
        return False
    if not build_command(job):
        return False
    if is_workflow_dep(job):
        return False
    if is_filewatcher(job):
        return False
    return True


def extract_rows(data: Dict[str, Any]) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []

    for folder_name, folder in data.items():
        if not isinstance(folder, dict):
            continue

        jobs = folder.get("Jobs")
        if not isinstance(jobs, list):
            continue

        for job in jobs:
            if not isinstance(job, dict):
                continue
            if not is_candidate_job(job):
                continue

            when = job.get("When")
            schedule = schedule_to_text(when)

            rows.append({
                "Folder": folder_name,
                "JobName": job.get("Name", "") or "",
                "Application": job.get("Application", "") or "",
                "SubApplication": job.get("SubApplication", "") or "",
                "Host": job.get("Host", "") or "",
                "RunAs": job.get("RunAs", "") or "",
                "CommandOrScript": build_command(job),
                "Schedule": schedule,
                "HasCalendar": "YES" if has_calendar(when) else "NO",
                "ScheduleMissing": "YES" if not schedule.strip() else "NO",
            })

    return rows


def write_csv(rows: List[Dict[str, str]], out_csv: Path) -> None:
    fieldnames = [
        "Folder", "JobName", "Application", "SubApplication", "Host", "RunAs",
        "CommandOrScript", "Schedule", "HasCalendar", "ScheduleMissing"
    ]
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)


def write_xlsx(rows: List[Dict[str, str]], out_xlsx: Path) -> None:
    headers = [
        "Folder", "JobName", "Application", "SubApplication", "Host", "RunAs",
        "CommandOrScript", "Schedule", "HasCalendar", "ScheduleMissing"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Standalone Jobs"
    ws.append(headers)

    # Header style
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = header_font

    # Row highlight style (red)
    red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")  # light red
    red_font = Font(color="FF9C0006")  # dark red

    for r_idx, r in enumerate(rows, start=2):
        ws.append([r.get(h, "") for h in headers])

        # Highlight entire row if Calendar exists OR Schedule missing
        if r.get("HasCalendar") == "YES" or r.get("ScheduleMissing") == "YES":
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=r_idx, column=c)
                cell.fill = red_fill
                cell.font = red_font

    # Auto width + freeze
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

    ws.freeze_panes = "A2"
    wb.save(out_xlsx)


def main():
    ap = argparse.ArgumentParser(
        description="Extract standalone Control-M jobs from JSON and auto-generate CSV/XLSX with same base filename."
    )
    ap.add_argument("input_json", help="Path to Control-M JSON export")
    ap.add_argument("--no-csv", action="store_true", help="Do not generate CSV")
    ap.add_argument("--no-xlsx", action="store_true", help="Do not generate XLSX")
    args = ap.parse_args()

    in_path = Path(args.input_json).expanduser().resolve()
    if not in_path.exists():
        raise SystemExit(f"[ERROR] File not found: {in_path}")

    base = in_path.with_suffix("")  # remove .json
    out_csv = base.with_suffix(".csv")
    out_xlsx = base.with_suffix(".xlsx")

    with in_path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    rows = extract_rows(data)

    if not args.no_csv:
        write_csv(rows, out_csv)

    if not args.no_xlsx:
        write_xlsx(rows, out_xlsx)

    produced = []
    if not args.no_csv:
        produced.append(str(out_csv))
    if not args.no_xlsx:
        produced.append(str(out_xlsx))

    print(f"[OK] Extracted {len(rows)} standalone jobs")
    for p in produced:
        print(f" - {p}")


if __name__ == "__main__":
    main()
